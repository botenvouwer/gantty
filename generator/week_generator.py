from datetime import datetime
from openpyxl.styles import Font
from pathlib import Path

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from generator.time_iterator import TimeIteratorMode, TimeIterator


def generate_week_gannt(months_duration, start_year=None, start_month=None, path=None, name='gantt_week_template'):
    start_year = start_year or datetime.datetime.now().year
    start_month = start_month or datetime.datetime.now().month
    mode = TimeIteratorMode.WEEKS
    path = path or (Path.home() / 'Documents' / (name + '.xlsx'))

    timeIterate = TimeIterator(start_year, start_month, months_duration, mode)

    matrix_h_start = 3
    matrix_v_start = 7
    matrix_h_end = len(timeIterate)
    matrix_v_end = 200
    matrix_color = 'BFBFBF'

    red = 'FF0000'

    italic = Font(italic=True)
    bold = Font(bold=True)
    small = Font(size=9)
    matrix_fill = PatternFill(patternType='solid', bgColor=red)
    left_align = Alignment(horizontal='left')
    center_align = Alignment(horizontal='center', vertical='center')
    middle_align = Alignment(horizontal='center')
    border = Border(bottom=Side(style='thin'))
    matrix_border = Border(bottom=Side(style='thin', color=matrix_color), top=Side(style='thin', color=matrix_color),
                           left=Side(style='thin', color=matrix_color), right=Side(style='thin', color=matrix_color))
    date_format = 'D-M-YYYY'

    styles = {}

    YEAR_EVEN = 'year_even'
    year_even = NamedStyle(name=YEAR_EVEN)
    year_even.font = Font(color='00FFFFFF')
    year_even.fill = PatternFill(patternType='solid', fgColor='808080')
    year_even.alignment = left_align
    styles[YEAR_EVEN] = year_even

    YEAR_ODD = 'year_odd'
    year_odd = NamedStyle(name=YEAR_ODD)
    year_odd.font = Font(color='00FFFFFF')
    year_odd.fill = PatternFill(patternType='solid', fgColor='595959')
    year_odd.alignment = left_align
    styles[YEAR_ODD] = year_even

    month_colors = (
    '9BBB59', '4BACC6', 'F79646', 'E0D60E', '8064A2', 'C0504D', '9BBB59', '4BACC6', 'F79646', 'E0D60E', '8064A2',
    'C0504D')
    month_day_colors = (
    'D8E4BC', 'B7DEE8', 'FCD5B4', 'F8F388', 'CCC0DA', 'E6B8B7', 'D8E4BC', 'B7DEE8', 'FCD5B4', 'F8F388', 'CCC0DA',
    'E6B8B7')

    y = 0
    for c in month_colors:
        y += 1
        l = 'month_' + str(y)
        month_s = NamedStyle(name=l)
        month_s.font = Font(color='00FFFFFF', italic=True)
        month_s.fill = PatternFill(fill_type='solid', fgColor=c)
        month_s.alignment = left_align

        styles[l] = month_s

    quarter_colors = ('D9D9D9', 'F2F2F2', 'D9D9D9', 'F2F2F2')

    y = 0
    for c in quarter_colors:
        y += 1
        l = 'quarter_' + str(y)
        quarter_s = NamedStyle(name=l)
        quarter_s.font = Font(bold=True, size=11)
        quarter_s.fill = PatternFill(fill_type='solid', fgColor=c)
        quarter_s.alignment = left_align

        styles[l] = quarter_s

    WEEK_EVEN = 'week_even'
    week_even = NamedStyle(name=WEEK_EVEN)
    week_even.font = Font(size=9)
    week_even.fill = PatternFill(patternType='solid', fgColor='F2F2F2')
    week_even.alignment = left_align
    styles[WEEK_EVEN] = year_even

    WEEK_ODD = 'week_odd'
    week_odd = NamedStyle(name=WEEK_ODD)
    week_odd.font = Font(size=9)
    week_odd.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
    week_odd.alignment = left_align
    styles[WEEK_ODD] = year_even

    DAY_EVEN = 'day_even'
    day_even = NamedStyle(name=DAY_EVEN)
    day_even.font = Font(size=9, italic=True)
    day_even.fill = PatternFill(patternType='solid', fgColor='F2F2F2')
    day_even.alignment = center_align
    styles[DAY_EVEN] = year_even

    DAY_ODD = 'day_odd'
    day_odd = NamedStyle(name=DAY_ODD)
    day_odd.font = Font(size=9, italic=True)
    day_odd.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
    day_odd.alignment = center_align
    styles[DAY_ODD] = year_even

    matrix_font = Font(size=16, color=red)

    MATRIX_CELL_EVEN = 'matrix_cell_even'
    matrix_cell_even = NamedStyle(name=MATRIX_CELL_EVEN)
    matrix_cell_even.font = matrix_font
    matrix_cell_even.border = matrix_border
    matrix_cell_even.fill = PatternFill(patternType='solid', fgColor='F2F2F2')
    matrix_cell_even.alignment = center_align
    styles[MATRIX_CELL_EVEN] = year_even

    MATRIX_CELL_ODD = 'matrix_cell_odd'
    matrix_cell_odd = NamedStyle(name=MATRIX_CELL_ODD)
    matrix_cell_odd.font = matrix_font
    matrix_cell_odd.border = matrix_border
    matrix_cell_odd.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
    matrix_cell_odd.alignment = center_align
    styles[MATRIX_CELL_ODD] = year_even

    y = 0
    for c in month_day_colors:
        y += 1
        l = 'month_day_' + str(y)
        month_s = NamedStyle(name=l)
        month_s.font = Font(size=9)
        month_s.fill = PatternFill(fill_type='solid', fgColor=c)
        month_s.alignment = center_align

        styles[l] = month_s

    wb = Workbook()
    worksheet: Worksheet = wb.active
    worksheet.title = name

    worksheet.column_dimensions['A'].width = 1.2
    worksheet.column_dimensions['B'].width = 43
    worksheet.column_dimensions['C'].width = 9

    # Take into consideration that the max cols in Excel are 16384
    max_columns_in_excel = matrix_h_end + matrix_h_start + 1
    for y in range(matrix_h_start + 1, max_columns_in_excel):
        worksheet.column_dimensions[get_column_letter(y)].width = 2.8

    # Prepare 200 rows to use as task -> this should be more than enough
    worksheet.row_dimensions[1].height = 3
    for y in range(2, matrix_v_end + 1):
        worksheet.row_dimensions[y].height = 15

    # Today value
    worksheet.cell(row=2, column=2).value = 'Vandaag'
    worksheet.cell(row=2, column=2).font = bold
    worksheet.cell(row=2, column=2).alignment = middle_align
    worksheet.cell(row=3, column=2).value = '=TODAY()'
    worksheet.cell(row=3, column=2).number_format = date_format
    worksheet.cell(row=3, column=2).alignment = middle_align

    # Current week value
    worksheet.cell(row=4, column=2).value = '=WEEKNUM(B3)&"-"&YEAR(B3)'
    worksheet.cell(row=4, column=2).alignment = middle_align

    worksheet.cell(row=6, column=2).value = 'Actie'
    worksheet.cell(row=6, column=2).font = bold
    worksheet.cell(row=6, column=3).value = 'Duur'
    worksheet.cell(row=6, column=3).font = bold

    for y in range(matrix_v_start, matrix_v_end + 1):
        f = f'=COUNTIF(D{y}:ZZ{y},"*")'
        worksheet.cell(row=y, column=3).value = f

    worksheet.conditional_formatting.add(f'$D${matrix_v_start}:$ZZ${matrix_v_end}',
                                         FormulaRule(formula=[f'D{matrix_v_start}<>""'], stopIfTrue=False,
                                                     fill=matrix_fill, font=matrix_font))

    # Highlight today
    border_highlight_top = Border(left=Side(style='thin', color=red), right=Side(style='thin', color=red),top=Side(style='thin', color=red))
    border_highlight = Border(left=Side(style='thin', color=red), right=Side(style='thin', color=red))
    worksheet.conditional_formatting.add(
        f'$D$5:$ZZ$5',
        FormulaRule(
            formula=[f'TEXT($B$4,"0")=TEXT(D$1,"0")'],
            stopIfTrue=False,
            border=border_highlight_top
        )
    )
    worksheet.conditional_formatting.add(
        f'$D$6:$ZZ$200',
        FormulaRule(
            formula=[f'TEXT($B$4,"0")=TEXT(D$1,"0")'],
            stopIfTrue=False,
            border=border_highlight
        )
    )

    handle_first_month = True
    last_month = 0
    first_week_number = 0
    last_week_number = 53 # so the first time we skip this
    first_year = 0
    last_year = 0
    weeks_in_month = 0
    year_row = 3
    last_quarter = 0
    weeks_in_quarter = 0
    handle_first_quarter = True
    for passed_months, year, days_left_in_year, month, month_name, days_in_month, y, week_number in timeIterate:
        column_i = matrix_h_start + y
        weeks_in_month += 1
        weeks_in_quarter += 1

        if first_week_number == 0:
            first_week_number = week_number
            first_year = year

        # Year
        if last_week_number > week_number:
            worksheet.cell(row=year_row, column=column_i).value = year
            worksheet.cell(row=year_row, column=column_i).style = year_even if year % 2 == 0 else year_odd

        # Quarter
        quarter = (month - 1) // 3 + 1  # Calculate quarter (1, 2, 3, or 4)
        row = 4
        worksheet.cell(row=row, column=column_i).value = f"Q{quarter}"
        quarter_style = styles['quarter_' + str(quarter)]
        worksheet.cell(row=row, column=column_i).style = quarter_style

        if last_quarter != 0 and last_quarter != quarter:
            # Merge cells for the previous quarter
            column_i_start = column_i - weeks_in_quarter + 1
            if handle_first_quarter:
                handle_first_quarter = False
                column_i_start = matrix_h_start + 1
            column_i_end = column_i - 1
            if column_i_start <= column_i_end:  # Ensure valid range for merging
                worksheet.merge_cells(start_row=row, end_row=row, start_column=column_i_start, end_column=column_i_end)
            weeks_in_quarter = 1  # Reset weeks in quarter

        # Month
        row = 5
        month_style = styles['month_' + str(month)]
        worksheet.cell(row=row, column=column_i).value = month_name
        worksheet.cell(row=row, column=column_i).style = month_style

        if last_month != 0 and last_month != month:
            # weeks_in_month = timeIterate.get_weeks_in_month(year, month)
            column_i_start = column_i - weeks_in_month + 1
            if handle_first_month:
                handle_first_month = False
                column_i_start = matrix_h_start + 1
            column_i_end = column_i - 1
            worksheet.merge_cells(start_row=row, end_row=row, start_column=column_i_start, end_column=column_i_end)
            weeks_in_month = 1

        last_month = month
        last_quarter = quarter

        # Week number
        row = 6

        week_is_even = week_number % 2 == 0
        worksheet.cell(row=row, column=column_i).value = week_number
        worksheet.cell(row=row, column=column_i).style = week_even if week_is_even else week_odd

        # Metadata -> used to highlight which day it is today
        row = 1

        worksheet.cell(row=row, column=column_i).value = f'{week_number}-{year}'
        worksheet.cell(row=row, column=column_i).number_format = date_format

        last_week_number = week_number
        last_year = year

    first_year_index_start = matrix_h_start + 1
    first_year_index_end = timeIterate.get_number_of_weeks_for_year(start_year) - first_week_number + first_year_index_start
    start_index_for_year = first_year_index_start

    # merge years
    for y in range(first_year, last_year + 1):

        match y:
            case _ if y == first_year:
                end_index_for_year = first_year_index_end
            case _ if y == last_year:
                start_index_for_year = end_index_for_year + 1
                end_index_for_year = start_index_for_year + last_week_number - 1
            case _:
                weeks_in_year = timeIterate.get_number_of_weeks_for_year(y)
                start_index_for_year = end_index_for_year + 1
                end_index_for_year = start_index_for_year + weeks_in_year - 1

        worksheet.merge_cells(start_row=year_row, start_column=start_index_for_year, end_row=year_row, end_column=end_index_for_year)

    # merge last quarter
    if weeks_in_quarter > 0:
        row = 4
        column_i_start = column_i - weeks_in_quarter + 1
        if handle_first_quarter:  # If only one quarter was processed
            column_i_start = matrix_h_start + 1
        column_i_end = column_i
        if column_i_start <= column_i_end:
            worksheet.merge_cells(start_row=row, end_row=row, start_column=column_i_start, end_column=column_i_end)

    # Merge last month
    if weeks_in_month > 0:
        row = 5
        column_i_start = column_i - weeks_in_month + 1
        if handle_first_month:  # If only one month was processed
            column_i_start = matrix_h_start + 1
        column_i_end = column_i
        if column_i_start <= column_i_end:
            worksheet.merge_cells(start_row=row, end_row=row, start_column=column_i_start, end_column=column_i_end)

    # Set border
    for y in range(1, max_columns_in_excel):
        worksheet.cell(row=6, column=y).border = border

    worksheet.freeze_panes = worksheet['D7']

    print(path)
    wb.save(filename=path)
