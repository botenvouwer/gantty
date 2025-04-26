import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from generator.time_iterator import TimeIteratorMode, TimeIterator


def generate_day_gannt(months_duration, start_year=None, start_month=None, path=None, workdays_only=True):
    start_year = start_year or datetime.datetime.now().year
    start_month = start_month or datetime.datetime.now().month
    mode = TimeIteratorMode.DAYS_NO_WEEKEND if workdays_only else TimeIteratorMode.DAYS
    name = f"gantt_{mode.name.lower()}_template"
    path = path or (Path.home() / 'Documents' / (name + '.xlsx'))

    time_iterate = TimeIterator(start_year, start_month, months_duration, mode)

    matrix_h_start = 3
    matrix_h_start_letter = get_column_letter(matrix_h_start + 1)
    matrix_v_start = 7
    matrix_h_end = len(time_iterate)
    matrix_v_end = 200
    matrix_color = 'BFBFBF'

    red = 'FF0000'

    italic = Font(italic=True)
    bold = Font(bold=True)
    small = Font(size=9)
    matrix_fill = PatternFill(patternType='solid', bgColor=red)
    left_align = Alignment(horizontal='left')
    center_align = Alignment(horizontal='center', vertical='center')
    middle_align = Alignment(horizontal='center')
    border = Border(bottom=Side(style='thin'))
    matrix_border = Border(bottom=Side(style='thin', color=matrix_color), top=Side(style='thin', color=matrix_color),
                           left=Side(style='thin', color=matrix_color), right=Side(style='thin', color=matrix_color))
    date_format = 'D-M-YYYY'

    styles = {}

    YEAR_EVEN = 'year_even'
    year_even = NamedStyle(name=YEAR_EVEN)
    year_even.font = Font(color='00FFFFFF')
    year_even.fill = PatternFill(patternType='solid', fgColor='808080')
    year_even.alignment = left_align
    styles[YEAR_EVEN] = year_even

    YEAR_ODD = 'year_odd'
    year_odd = NamedStyle(name=YEAR_ODD)
    year_odd.font = Font(color='00FFFFFF')
    year_odd.fill = PatternFill(patternType='solid', fgColor='595959')
    year_odd.alignment = left_align
    styles[YEAR_ODD] = year_even

    month_colors = (
    '9BBB59', '4BACC6', 'F79646', 'E0D60E', '8064A2', 'C0504D', '9BBB59', '4BACC6', 'F79646', 'E0D60E', '8064A2',
    'C0504D')
    month_day_colors = (
    'D8E4BC', 'B7DEE8', 'FCD5B4', 'F8F388', 'CCC0DA', 'E6B8B7', 'D8E4BC', 'B7DEE8', 'FCD5B4', 'F8F388', 'CCC0DA',
    'E6B8B7')

    i = 0
    for c in month_colors:
        i += 1
        l = 'month_' + str(i)
        month_s = NamedStyle(name=l)
        month_s.font = Font(color='00FFFFFF', italic=True)
        month_s.fill = PatternFill(fill_type='solid', fgColor=c)
        month_s.alignment = left_align

        styles[l] = month_s

    WEEK_EVEN = 'week_even'
    week_even = NamedStyle(name=WEEK_EVEN)
    week_even.font = Font(size=9)
    week_even.fill = PatternFill(patternType='solid', fgColor='F2F2F2')
    week_even.alignment = left_align
    styles[WEEK_EVEN] = year_even

    WEEK_ODD = 'week_odd'
    week_odd = NamedStyle(name=WEEK_ODD)
    week_odd.font = Font(size=9)
    week_odd.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
    week_odd.alignment = left_align
    styles[WEEK_ODD] = year_even

    DAY_EVEN = 'day_even'
    day_even = NamedStyle(name=DAY_EVEN)
    day_even.font = Font(size=9, italic=True)
    day_even.fill = PatternFill(patternType='solid', fgColor='F2F2F2')
    day_even.alignment = center_align
    styles[DAY_EVEN] = year_even

    DAY_ODD = 'day_odd'
    day_odd = NamedStyle(name=DAY_ODD)
    day_odd.font = Font(size=9, italic=True)
    day_odd.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
    day_odd.alignment = center_align
    styles[DAY_ODD] = year_even

    matrix_font = Font(size=16, color=red)

    MATRIX_CELL_EVEN = 'matrix_cell_even'
    matrix_cell_even = NamedStyle(name=MATRIX_CELL_EVEN)
    matrix_cell_even.font = matrix_font
    matrix_cell_even.border = matrix_border
    matrix_cell_even.fill = PatternFill(patternType='solid', fgColor='F2F2F2')
    matrix_cell_even.alignment = center_align
    styles[MATRIX_CELL_EVEN] = year_even

    MATRIX_CELL_ODD = 'matrix_cell_odd'
    matrix_cell_odd = NamedStyle(name=MATRIX_CELL_ODD)
    matrix_cell_odd.font = matrix_font
    matrix_cell_odd.border = matrix_border
    matrix_cell_odd.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
    matrix_cell_odd.alignment = center_align
    styles[MATRIX_CELL_ODD] = year_even

    i = 0
    for c in month_day_colors:
        i += 1
        l = 'month_day_' + str(i)
        month_s = NamedStyle(name=l)
        month_s.font = Font(size=9)
        month_s.fill = PatternFill(fill_type='solid', fgColor=c)
        month_s.alignment = center_align

        styles[l] = month_s

    wb = Workbook()
    worksheet: Worksheet = wb.active
    worksheet.title = name

    worksheet.column_dimensions['A'].width = 1.2
    worksheet.column_dimensions['B'].width = 43
    worksheet.column_dimensions['C'].width = 9

    # Take into consideration that the max cols in Excel are 16384
    max_columns_in_excel = matrix_h_end + matrix_h_start + 1
    for i in range(matrix_h_start + 1, max_columns_in_excel):
        worksheet.column_dimensions[get_column_letter(i)].width = 2.8

    # Prepare 200 rows to use as task -> this should be more than enough
    worksheet.row_dimensions[1].height = 3
    for i in range(2, matrix_v_end + 1):
        worksheet.row_dimensions[i].height = 15

    worksheet.cell(row=2, column=2).value = 'Vandaag'
    worksheet.cell(row=2, column=2).font = bold
    worksheet.cell(row=2, column=2).alignment = middle_align
    worksheet.cell(row=3, column=2).value = '=TODAY()'
    worksheet.cell(row=3, column=2).number_format = date_format
    worksheet.cell(row=3, column=2).alignment = middle_align

    worksheet.cell(row=6, column=2).value = 'Actie'
    worksheet.cell(row=6, column=2).font = bold
    worksheet.cell(row=6, column=3).value = 'Duur'
    worksheet.cell(row=6, column=3).font = bold

    for i in range(matrix_v_start, matrix_v_end + 1):
        f = f'=COUNTIF(D{i}:ZZ{i},"*")'
        worksheet.cell(row=i, column=3).value = f

    worksheet.conditional_formatting.add(f'$D${matrix_v_start}:$ZZ${matrix_v_end}',
                                         FormulaRule(formula=[f'D{matrix_v_start}<>""'], stopIfTrue=False,
                                                     fill=matrix_fill, font=matrix_font))

    # Highlight today
    border_highlight_top = Border(left=Side(style='thin', color=red), right=Side(style='thin', color=red),
                                  top=Side(style='thin', color=red))
    border_highlight = Border(left=Side(style='thin', color=red), right=Side(style='thin', color=red))
    worksheet.conditional_formatting.add(f'$D$5:$ZZ$5',
                                         FormulaRule(formula=[f'TEXT($B$3,"0")=TEXT(D$1,"0")'], stopIfTrue=False,
                                                     border=border_highlight_top))
    worksheet.conditional_formatting.add(f'$D$6:$ZZ$200',
                                         FormulaRule(formula=[f'TEXT($B$3,"0")=TEXT(D$1,"0")'], stopIfTrue=False,
                                                     border=border_highlight))

    days_in_week = mode.value
    sum_days_in_year = 1
    sum_days_in_month = 1
    for passed_months, year, days_left_in_year, month, month_name, days_in_month, i, week_number, day_number, day, day_name in time_iterate:
        column_i = matrix_h_start + i

        # Year
        if i % sum_days_in_year == 0:
            row = 2
            worksheet.cell(row=row, column=column_i).value = year
            worksheet.cell(row=row, column=column_i).style = year_even if year % 2 == 0 else year_odd
            worksheet.merge_cells(start_row=row, start_column=column_i,
                                  end_row=row, end_column=column_i + days_left_in_year - 1)

            sum_days_in_year += days_left_in_year

        # Month
        if i % sum_days_in_month == 0:
            row = 3
            month_style = styles['month_' + str(month)]
            worksheet.cell(row=row, column=column_i).value = month_name
            worksheet.cell(row=row, column=column_i).style = month_style
            worksheet.merge_cells(start_row=row, start_column=column_i,
                                  end_row=row, end_column=column_i + days_in_month - 1)

            sum_days_in_month += days_in_month

        week_is_even = week_number % 2 == 0

        # Week number
        if i == 1 or day == 0:
            row = 4

            end_column_i = column_i + days_in_week - 1

            if i == 1:
                end_column_i = column_i + days_in_week - day - 1

            worksheet.cell(row=row, column=column_i).value = week_number
            worksheet.cell(row=row, column=column_i).style = week_even if week_is_even else week_odd
            worksheet.merge_cells(start_row=row, start_column=column_i, end_row=row, end_column=end_column_i)

        # Day
        row = 5
        worksheet.cell(row=row, column=column_i).value = day_name
        worksheet.cell(row=row, column=column_i).style = week_even if week_is_even else week_odd

        # Day of month
        row = 6
        month_day_style = styles['month_day_' + str(month)]
        worksheet.cell(row=row, column=column_i).value = day_number
        worksheet.cell(row=row, column=column_i).style = month_day_style

        # Metadata -> used to highlight which day it is today
        worksheet.cell(row=1, column=column_i).value = f'{day_number}-{month}-{year}'
        worksheet.cell(row=1, column=column_i).number_format = date_format

        # matrix row styling
        for row_i in range(matrix_v_start, matrix_v_end + 1):
            worksheet.cell(row=row_i, column=column_i).style = matrix_cell_even if week_is_even else matrix_cell_odd

    for i in range(1, max_columns_in_excel):
        worksheet.cell(row=6, column=i).border = border

    # GET START LETTER

    c = worksheet[f"{matrix_h_start_letter}{matrix_v_start}"]
    worksheet.freeze_panes = c

    print(f"Saved to: {path}")
    wb.save(filename=path)
