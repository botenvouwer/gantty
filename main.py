import datetime
from math import floor
from pathlib import Path
import calendar

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from time_util import days_left_in_year

name = "gantt"
path = Path.home() / (name + '.xlsx')
# path = 'custom/path'

italic = Font(italic=True)
bold = Font(bold=True)
small = Font(size=9)
left_align = Alignment(horizontal='left')
center_align = Alignment(horizontal='center', vertical='center')

border = Border(bottom=Side(style='thin'))

wb = Workbook()
worksheet: Worksheet = wb.active
worksheet.title = name

worksheet.column_dimensions['A'].width = 1.2
worksheet.column_dimensions['B'].width = 43
worksheet.column_dimensions['C'].width = 9

max_columns_in_excel = 16384
for i in range(1, 16384):
    if i >= 4:
        worksheet.column_dimensions[get_column_letter(i)].width = 2.8

    worksheet.row_dimensions[i].height = 15
    worksheet.cell(row=6, column=i).border = border

worksheet.cell(row=6, column=2).value = 'Actie'
worksheet.cell(row=6, column=2).font = bold

day_names = ('ma', 'di', 'wo', 'do', 'vr', 'za', 'zo')
month_names = ('Januari', 'Februari', 'Maart', 'April', 'Mei', 'Juni', 'Juli', 'Augustus', 'Sebtember', 'Oktober', 'November', 'December')
c = calendar.Calendar()

start_year = 2022
start_month = 8
months_duration = 12

end_year = start_year + floor((months_duration - start_month) / 12) + 1
end_month = (months_duration - (12 - start_month)) % 12 - 1

ym_start = 12 * start_year + start_month - 1
ym_end = 12 * end_year + end_month
i = 0
ii = 0
matrix_start = 3
week_start_index = 4
week_start_end_index = 10
days_in_month = calendar.monthrange(start_year, start_month)[1]
days_in_year = days_left_in_year(start_year, start_month)
sum_year_thing = 1
for ym in range(ym_start, ym_end):
    y, m = divmod(ym, 12)
    year = y
    month = m + 1
    i += 1

    for day_number in c.itermonthdays(year, month):
        if day_number == 0:
            continue

        ii += 1
        day = calendar.weekday(year, month, day_number)
        week_number = datetime.date(year, month, day_number).isocalendar().week

        if ii % sum_year_thing == 0:
            worksheet.cell(row=2, column=ii + matrix_start).value = year
            worksheet.cell(row=2, column=ii + matrix_start).alignment = left_align
            worksheet.merge_cells(start_row=2, start_column=ii + matrix_start,
                                  end_row=2, end_column=ii + matrix_start + days_in_year - 1)

            next_year = year + 1
            sum_year_thing += days_in_year
            days_in_year = days_left_in_year(next_year, end_month=end_month) if end_year == next_year else days_left_in_year(next_year)

        if day_number == 1:
            worksheet.cell(row=3, column=ii + matrix_start).value = month_names[month - 1]
            worksheet.cell(row=3, column=ii + matrix_start).alignment = left_align
            worksheet.merge_cells(start_row=3, start_column=ii + matrix_start,
                                  end_row=3, end_column=ii + matrix_start + days_in_month - 1)

            next_month = calendar._nextmonth(year=year, month=month)[1]
            days_in_month = calendar.monthrange(year, next_month)[1] if next_month > 1 else calendar.monthrange(year + 1, next_month)[1]

        if ii % 7 == 1:
            worksheet.cell(row=4, column=week_start_index).value = week_number
            worksheet.cell(row=4, column=week_start_index).alignment = left_align
            worksheet.merge_cells(start_row=4, start_column=week_start_index, end_row=4, end_column=week_start_end_index)
            week_start_index = week_start_index + 7
            week_start_end_index = week_start_end_index + 7

        worksheet.cell(row=5, column=ii + matrix_start).value = day_names[day]
        worksheet.cell(row=5, column=ii + matrix_start).alignment = center_align
        worksheet.cell(row=5, column=ii + matrix_start).font = small
        worksheet.cell(row=6, column=ii + matrix_start).value = day_number
        worksheet.cell(row=6, column=ii + matrix_start).alignment = center_align
        worksheet.cell(row=6, column=ii + matrix_start).font = small

        # print(i, year, month, month_names[month - 1], day_number, week_number, day, day_names[day])

print(path)
wb.save(filename=path)
